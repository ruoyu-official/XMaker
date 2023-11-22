from threading import Event, Thread
from queue import Queue
from openpyxl import Workbook, load_workbook
from utils.config import *
from utils.lib import *
from utils.progress import *

CURDIR = CWD()  # 当前路径
_OutFile_Maxlen = 0  # 输出文件名最大长度

# 计算过程所需各值
_Y_index = _Sg_index = _Vars_len = _Y_min = _Sg_max = _Sg_sum = 0

_TimeEV = Event()
_ParamQue = Queue()  # IN消息队列
_ResLine = ""  # OUT消息


def handle_calc(y=0, sg=0):
    """
    计算
    """
    if sg <= 0:
        return
    global _Y_min, _Sg_max, _Sg_sum
    # 初始化
    if not (_Y_min or _Sg_max):
        _Y_min, _Sg_max = y, sg
    # 计算
    if y < _Y_min:
        _Y_min = y
    if sg > _Sg_max:
        _Sg_max = sg
    if sg < 0.1:
        _Sg_sum += sg


def handle_load(front: str, out_path: str):
    """
    读取已生成文件
    """
    start = NOW()
    global _ParamQue
    _ParamQue.put((front, start))  # 输入队列，存值
    _TimeThread = Thread(
        target=handle_time,
        kwargs={
            "thread_ev": _TimeEV,
            "params_queue": _ParamQue,
        },
    )
    handle_thread(_TimeThread)
    wb = None
    try:
        if not out_path:
            return ""
        wb = load_workbook(out_path, read_only=True)
        ws = wb.active
        _d = ws.calculate_dimension(force=True)  # 区域范围 A1:AO30002
        _last = _d.split(":")[1]
        _last = int("".join(c for c in _last if c.isdigit()))
        global _Y_min, _Sg_max, _Sg_sum
        _Y_min, _Sg_max, _Sg_sum = [item.value for item in ws[_last] if item.value]
        # print(_temp)
        _values = (
            numstr(_Y_min, precision=PRECS),
            numstr(_Sg_max, precision=PRECS),
            numstr(_Sg_sum, precision=PRECS),
        )
        _Y_min = _Sg_max = _Sg_sum = 0
    except (KeyboardInterrupt, PermissionError, ValueError, Exception) as e:
        # print(str(e))
        handle_thread(_TimeThread, stop=True)
        global _ResLine
        PRINT(
            front,
            res=f"{FRONT['calc']}{ERR['errupt']}",
            ok=True,
            cover=natlen(_ResLine),
        )
        _ResLine = ""
        _Y_min = _Sg_max = _Sg_sum = 0
        if wb:
            wb.close()
        return ""
    handle_thread(_TimeThread, stop=True)
    wb.close()
    return f"{FRONTED['calc']}{zipstr(CVAR,_values)}"


def handle_time(thread_ev: Event, params_queue: Queue, write=False):
    """
    计时器
    """
    global _ResLine
    if not write:
        front, start = params_queue.get()  # 输入队列，取值
        progress = 0
        modlen = CALC_LEN + 1
        while not thread_ev.is_set():
            load = PROGRESS("calc", start, progress=progress)
            _ResLine = PRINT(front, f"{OK} {load}")
            progress = (progress + 1) % modlen
            SLEEP(CALC_FREQ)
    else:
        front, make, start = params_queue.get()  # 输入队列，取值
        progress = 0
        modlen = SAVE_LEN + 1
        while not thread_ev.is_set():
            save = PROGRESS("save", start, progress=progress)
            _ResLine = PRINT(front, make, save)
            progress = (progress + 1) % modlen
            SLEEP(SAVE_FREQ)


def handle_thread(thread: Thread, event=_TimeEV, stop=False):
    if not stop:
        event.clear()
        thread.start()
    else:
        event.set()
        if thread and thread.is_alive():
            thread.join()  # 等待线程结束


def handle_save(front: str, make: str, out_path: str, workbook: Workbook):
    """
    保存生成文件
    """
    start = NOW()  # 获取开始时间
    global _ParamQue
    _ParamQue.put((front, make, start))  # 输入队列，存值
    _TimeThread = Thread(  # 创建线程，绑定计时器，线程使用是一次性的
        target=handle_time,
        kwargs={
            "thread_ev": _TimeEV,
            "params_queue": _ParamQue,
            "write": True,
        },
    )
    handle_thread(_TimeThread)  # 计时线程启动
    # _TimeEV.clear()  # 清表，计时归零
    # _TimeThread.start()  # 计时开始
    global _Y_min, _Sg_max, _Sg_sum
    try:
        workbook.save(out_path)  # 保存
    except (KeyboardInterrupt, PermissionError, ValueError, Exception) as e:
        # print(f"保存异常")
        handle_thread(_TimeThread, stop=True)  # 线程结束
        global _ResLine
        PRINT(
            front,
            err=f"{FRONT['save']}{ERR['errupt']}，{FRONT['file']}{ERR['broken']}，{ERR['removed']}",
            cover=natlen(_ResLine),
        )
        _ResLine = ""  # 重置resline
        _Y_min = _Sg_max = _Sg_sum = 0
        workbook.close()  # 保存异常，关闭工作簿
        return ""
    _Y_min = _Sg_max = _Sg_sum = 0
    handle_thread(_TimeThread, stop=True)
    workbook.close()  # 正常关闭工作簿
    return PROGRESS("save", start, ok=True)


def handle_line(worksheet, row: int, line):
    """
    处理单行内容
    """
    _line = line
    # try:
    if row > 2:  # 单元处理
        for col, field in enumerate(_line):
            try:
                _line[col] = float(field)
            except ValueError:  # 无法转数字的单元值默认为0
                _line[col] = 0
    worksheet.append(_line)
    return _line  # 返回float数据


def handle_make(front: str, src_path: str, workbook: Workbook, worksheet):
    """
    生成、计算
    """
    start = NOW()  # 计时开始
    content = CONTENT(src_path)  # 获取文件内容

    maxlen = len(content)  # 全长
    checklen = round(maxlen / MAKE_LEN)  # 间隔

    tmp_line_str = ""
    global _Y_min, _Sg_max, _Sg_sum, _Y_index, _Sg_index, _Vars_len
    try:
        for row, line in enumerate(content, start=1):  # 行处理
            if row == 1:  # 表头
                sline = line.strip().split("=")[1].split()
                handle_line(worksheet, row, sline)

                if not _Vars_len:  # 初始化目标脚标
                    _Vars_len = len(sline)
                    _Y_index = sline.index("Y")
                    _Sg_index = sline.index("SG")

            elif row == 2:  # row==2 默认为空
                pass
                # handle_line(worksheet, row, [])
            elif row > 2:
                sline = line.strip().split()
                fline = handle_line(worksheet, row, sline)  # 写入行数据
                # 计算
                _y, _sg = (fline[_Y_index], fline[_Sg_index])
                handle_calc(_y, _sg)
            if row == 1 or row % checklen == 0:  # 设置进度刷新点
                progress = round(row / checklen)
                tmp_line_str = PRINT(front, PROGRESS("make", start, progress=progress))
        calc_line = [""] * _Vars_len
        calc_line[:3] = [_Y_min, _Sg_max, _Sg_sum]  # 计算结果
        handle_line(worksheet, 2, calc_line)  # 写入表格最后一行
        return PROGRESS("make", start, ok=True)  # 返回转换成功文本
    except (KeyboardInterrupt, PermissionError, Exception) as e:
        # print(str(e))
        PRINT(front, err=f"{FRONT['make']}{ERR['errupt']}", cover=natlen(tmp_line_str))
        _Y_min = _Sg_max = _Sg_sum = 0
        workbook.close()  # 转换异常关闭
        return ""


def handle_file(src_dir: str, src_file: str, first=False, force=False, calc=False):
    """
    处理文件
    """
    ext = EXTENSION(src_file)  # 获取扩展名
    if not ext in INEXT:  # 默认仅处理支持的格式文件
        return
    # 处理表页
    out_file = attach(src_file, attach_str=OUTEXT)  # 目标文件名（此后src_file可改）
    out_path = PJOIN(src_dir, out_file)  # 目标路径（此后out_file可改）

    _front = FRONTED["file"] if first else FRONTED["=file"]  # 头文
    _front_ok = _front + align(out_file, _OutFile_Maxlen, right=True)  # 完成名
    _ok = EXISTS(out_path)  # 获取存在状态
    _opened = XOPENED(src_dir, out_file)  # 获取打开状态
    global _Y_min, _Sg_max, _Sg_sum  # 申请全局
    if (not force) and _ok:  # 非强制模式下、已存在则直接打印  不执行操作
        # 读取表格 最后一行获取计算结果
        calc_str = ""
        if calc:
            # 读文件获取计算值
            calc_str = handle_load(front=_front_ok, out_path=out_path)
            if not calc_str:
                return
        PRINT(_front_ok, res=calc_str, ok=True)
        return

    # ext_filler = ANY if _ok else UDL  # 尾缀填充符 ，_表示目标文件不存在、*表示已存在但是强制转换
    _front_nok = _front + align(
        attach(src_file, attach(DOT, ANY, count=len(OUTEXT) - 1)),
        _OutFile_Maxlen,
        right=True,
    )  # 进行名
    if _opened:  # 打开状态下、如果转换，则报错而不执行
        PRINT(
            _front_ok,
            err=f"{FRONT['make']}{ERR['errupt']}，{FRONT['file']}{ERR['opened']}",
        )
        return
    # 新建工作簿，只读模式下保存略快
    wb = Workbook(write_only=True)
    # _datetime = DATETIME()
    ws = wb.create_sheet(title=src_file)

    make_str = handle_make(_front_nok, PJOIN(src_dir, src_file), wb, ws)  # 转化

    if not make_str:  # 转化异常
        return
    calc_str = ""
    if calc:  # 生成式结果
        _values = (
            numstr(_Y_min, precision=PRECS),
            numstr(_Sg_max, precision=PRECS),
            numstr(_Sg_sum, precision=PRECS),
        )
        _Y_min = _Sg_max = _Sg_sum = 0  # 重置
        calc_str = f"{FRONTED['calc']}{zipstr(CVAR,_values)}"
    save_str = handle_save(_front_nok, make_str, out_path, wb)  # 保存
    if not save_str:  # 保存异常
        if not _opened:  # 若目标文件本就不存在，则删除
            REMOVE(out_path)
        return
    ok_line = PRINT(_front_ok, make_str, save_str)  # 满进度
    # print("line length:", natlen(ok_line))
    PRINT(_front_ok, ok=True, res=calc_str, cover=natlen(ok_line))  # 打印成功默认信息


def handle_args(**kwargs):
    """
    处理指令
    """

    _make_path, _force, _calc = tuple(kwargs.values())  # 字典值转元组
    _default = not _make_path  # 路径为空则设置默认模式
    path = _make_path if _make_path else DEFAULT  # 如果为None就置为默认文件名
    # 路径预处理
    if not ISABS(path):  # 相对路径则补全（默认为当前路径）
        path = PJOIN(CURDIR, path)
    _exist = EXISTS(path)  # 目标路径是否有效
    _ptype = PTYPE(path)  # 目标路径类型
    _fronted_path = f"{FRONTED[_ptype]}<{path}>"
    if not _exist:  # 检查路径
        if not _default:
            print(f"{_fronted_path} {ERR['not_exist']}")
            return
        else:
            MAKEDIR(path)  # 保证默认文件夹存在
            _ptype = "dir"  # 类型指定为dir
            _fronted_path = f"{FRONTED[_ptype]}<{path}>"
    # return

    # 解析路径目标
    _dir, _files = PSRCS(path, _ptype)
    if not _files:  # 没有目标文件  退出
        pre = ERR["no"] if _ptype == "dir" else ERR["not"]
        _not_tar_file = f"{pre}{str(INEXT)}{FRONT['file']}"
        print(f"{_fronted_path} {_not_tar_file}")
        return
    print(_fronted_path)  # 打印路径信息
    global _OutFile_Maxlen  # 获取目标文件名最大长度
    _OutFile_Maxlen = maxlen(_files) + len(OUTEXT)

    # 根据预处理路径执行文件
    for index, _file in enumerate(_files):
        handle_file(_dir, _file, first=(index == 0), force=_force, calc=_calc)

    EXIT()  # 强制退出，避免调用atexit退出处理器
